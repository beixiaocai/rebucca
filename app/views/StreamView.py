# 作者：北小菜
# 官网：https://www.yuturuishi.com
# 微信：bilibili_bxc
# 哔哩哔哩主页：https://space.bilibili.com/487906612
# gitee地址：https://gitee.com/Vanishi/rebucca
# github地址：https://github.com/beixiaocai/rebucca
import json
from app.views.ViewsBase import *
from app.models import *
from django.shortcuts import render, redirect
from app.utils.Utils import buildPageLabels, group_by_field, GB28181CodeUtils
from app.utils.UploadUtils import UploadUtils
from app.utils.LogUtils import LogUtils

def online(request):
    context = {}
    # data = Camera.objects.all().order_by("-sort")
    return render(request, 'app/stream/online.html', context)
def index(request):
    context = {}
    params = f_parseGetParams(request)
    stream_code = params.get("code","").strip()
    context["search_text"] = stream_code # v4.716新增
    return render(request, 'app/stream/index.html', context)

def api_openAddContext(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    request_ip = f_parseRequestIp(request)  # （v4.725新增） 获取实时请求ip
    stream_info = {}

    if "GET" == request.method:
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parseGetParams(request)
            g_logger.info("StreamView.openAddContext() request_ip:%s,params:%s" % (request_ip,str(params)))

            pull_stream_username = ""
            pull_stream_password = ""
            stream = g_database.select("select id,pull_stream_username,pull_stream_password from av_stream where pull_stream_type=1 order by id desc limit 1")
            if len(stream) > 0:
                pull_stream_username = stream[0]["pull_stream_username"]
                pull_stream_password = stream[0]["pull_stream_password"]


            code = GB28181CodeUtils().generate_by_time()
            app = g_zlm.default_stream_app
            name = code
            camera_device_id = "default"

            stream_info = {
                "code": code,
                "app": app,
                "name": name,
                "camera_device_id": camera_device_id,
                "is_audio": 0,
                "pull_stream_username": pull_stream_username,
                "pull_stream_password": pull_stream_password,
                "rtspUrl": g_zlm.get_rtspUrl(app=app,name=name,request_ip=request_ip),
                "hlsUrl": g_zlm.get_hlsUrl(app=app,name=name,request_ip=request_ip),
                "httpMp4Url": g_zlm.get_httpMp4Url(app=app,name=name,request_ip=request_ip),
                "wsMp4Url": g_zlm.get_wsMp4Url(app=app,name=name,request_ip=request_ip),
                "record_enable": 0,
            }
            ret = True
            msg = LANG_VIEWS_T(request, "msg_success")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg,
        "stream": stream_info,
        "pull_stream_types": g_pull_stream_types,
        "audio_types": get_audio_types(f_parseRequestLang(request))
    }
    g_logger.info("StreamView.openAddContext() res:%s" % str(res))
    return f_responseJson(res)

def api_openAdd(request):
    __ret = False
    __msg = LANG_VIEWS_T(request, "msg_unknown_error")

    if "POST" == request.method:
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            try:
                params = f_parsePostParams(request)
                lang = f_parseRequestLang(request)
                g_logger.info("StreamView.openAdd() params:%s" % str(params))

                code = params.get("code", "").strip()
                # app = params.get("app", "live").strip()
                # name = params.get("name", "").strip()
                nickname = params.get("nickname", "").strip()
                pull_stream_type = int(params.get("pull_stream_type", 0))
                pull_stream_url = params.get("pull_stream_url", "").strip()
                pull_stream_ip = params.get("pull_stream_ip", "").strip()
                pull_stream_port = int(params.get("pull_stream_port"))

                camera_name = params.get("camera_name", "").strip()
                camera_manufacturer = params.get("camera_manufacturer", "").strip()
                camera_device_id = params.get("camera_device_id", "").strip()

                remark = params.get("remark", "").strip()

                onvif_username = params.get("onvif_username", "").strip()
                onvif_password = params.get("onvif_password", "").strip()
                is_audio = int(params.get("is_audio", 0))
                record_enable = int(params.get("record_enable", 0))

                if code == "":
                    raise Exception(LANG_VIEWS_T(request, "stream_code_invalid"))


                stream = StreamModel.objects.filter(code=code).filter()
                if stream:
                    raise Exception(LANG_VIEWS_T(request, "msg_code_already_exists"))
                # if app == "":
                #     raise Exception(LANG_VIEWS_T(request, "stream_app_invalid"))
                # if name == "":
                #     raise Exception(LANG_VIEWS_T(request, "stream_name_invalid"))
                if nickname == "":
                    raise Exception(LANG_VIEWS_T(request, "stream_name_incorrect"))
                else:
                    nickname_len = len(nickname)
                    if len(nickname) > 50:
                        raise Exception(LANG_VIEWS_T(request, "stream_name_max_length") % nickname_len)

                if pull_stream_type == 21:
                    raise Exception(LANG_VIEWS_T(request, "stream_gb28181_hint"))
                else:
                    if pull_stream_type == 1 or pull_stream_type == 2 or pull_stream_type == 3 or pull_stream_type == 4:
                        pass
                    else:
                        raise Exception(LANG_VIEWS_T(request, "stream_protocol_not_supported"))

                if pull_stream_url == "":
                    raise Exception(LANG_VIEWS_T(request, "stream_url_incorrect"))

                if camera_device_id == "":
                    camera_device_id = code

                user_id = f_sessionReadUserId(request)
                now_date = datetime.now()

                stream = StreamModel()
                stream.user_id = user_id
                stream.sort = 0
                stream.code = code
                stream.app = APP_NAME_LIVE
                stream.name = code
                stream.pull_stream_url = pull_stream_url
                stream.pull_stream_type = pull_stream_type  # v4.405新增
                stream.pull_stream_transfer_mode = 0  # v4.405新增
                stream.pull_stream_ip = pull_stream_ip  # v4.405新增
                stream.pull_stream_port = pull_stream_port  # v4.405新增
                stream.pull_stream_username = onvif_username  # 当rtsp类型接入时，用于通过onvif探测接入的摄像头才会有username
                stream.pull_stream_password = onvif_password

                stream.nickname = nickname
                stream.remark = remark
                stream.forward_state = 0  # 默认未开启转发
                stream.is_audio = is_audio  # v4.405新增

                stream.camera_name = camera_name
                stream.camera_manufacturer = camera_manufacturer
                stream.camera_device_id = camera_device_id

                stream.record_enable = record_enable

                stream.create_time = now_date
                stream.last_update_time = now_date
                stream.add_type = 0
                stream.state = 0

                stream.save()

                # 添加日志
                user_id = f_sessionReadUserId(request)
                lang = f_parseRequestLang(request)
                LogUtils.add_stream_log(user_id=user_id, stream_code=code,
                                        log_type=LogUtils.LOG_TYPE_ADD, lang=lang)

                __msg = LANG_VIEWS_T(request, "msg_add_success")
                __ret = True

            except Exception as e:
                __msg = str(e)
        else:
            __msg = __check_msg
    else:
        __msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if __ret else 0,
        "msg": __msg
    }
    g_logger.info("StreamView.openAdd() res:%s" % str(res))
    return f_responseJson(res)

def api_openEditContext(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    request_ip = f_parseRequestIp(request)  # （v4.725新增） 获取实时请求ip
    stream_info = {}

    if "GET" == request.method:
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parseGetParams(request)
            g_logger.info("StreamView.openEditContext() request_ip:%s,params:%s" % (request_ip,str(params)))

            stream_code = params.get("code", "").strip()

            stream = g_database.select("select * from av_stream where code='%s' limit 1" % stream_code)
            if len(stream) > 0:
                stream_info = stream[0]
                camera_device_id = stream_info["camera_device_id"]
                if camera_device_id is None or camera_device_id == "":
                    stream_info["camera_device_id"] = stream_code

                ret = True
                msg = LANG_VIEWS_T(request, "msg_success")
            else:
                msg = LANG_VIEWS_T(request, "stream_not_found")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg,
        "stream": stream_info,
        "pull_stream_types": g_pull_stream_types,
        "audio_types": get_audio_types(f_parseRequestLang(request))
    }
    g_logger.info("StreamView.openEditContext() res:%s" % str(res))
    return f_responseJson(res)

def api_openStreamByAppAndName(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    data = {}

    if request.method == 'GET':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parseGetParams(request)
            g_logger.info("StreamView.openStreamByAppAndName() params:%s" % str(params))

            app = params.get("app", "").strip()
            name = params.get("name", "").strip()

            if app and name:
                streams = g_database.select("select * from av_stream where app='%s' and name='%s' limit 1" % (app, name))
                if len(streams) > 0:
                    data = streams[0]
                    ret = True
                    msg = LANG_VIEWS_T(request, "msg_success")
                else:
                    msg = LANG_VIEWS_T(request, "stream_not_found")
            else:
                msg = LANG_VIEWS_T(request, "stream_app_name_required")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg,
        "data": data
    }
    g_logger.info("StreamView.openStreamByAppAndName() res:%s" % str(res))
    return f_responseJson(res)

def api_openEdit(request):
    __ret = False
    __msg = LANG_VIEWS_T(request, "msg_unknown_error")

    if "POST" == request.method:
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            try:
                params = f_parsePostParams(request)
                g_logger.info("StreamView.openEdit() params:%s" % str(params))

                lang = f_parseRequestLang(request)

                code = params.get("code", "").strip()
                nickname = params.get("nickname", "").strip()
                pull_stream_url = params.get("pull_stream_url", "").strip()
                pull_stream_ip = params.get("pull_stream_ip", "").strip()
                pull_stream_port = int(params.get("pull_stream_port", 0))

                camera_name = params.get("camera_name", "").strip()
                camera_manufacturer = params.get("camera_manufacturer", "").strip()
                camera_device_id = params.get("camera_device_id", "").strip()

                remark = params.get("remark", "").strip()
                is_audio = int(params.get("is_audio", 0))
                record_enable = int(params.get("record_enable", 0))

                if code == "":
                    raise Exception(LANG_VIEWS_T(request, "stream_code_invalid"))

                if nickname == "":
                    raise Exception(LANG_VIEWS_T(request, "stream_name_incorrect"))
                else:
                    nickname_len = len(nickname)
                    if len(nickname) > 50:
                        raise Exception(LANG_VIEWS_T(request, "stream_name_max_length") % nickname_len)

                stream = StreamModel.objects.filter(code=code).first()
                if stream:
                    if (stream.pull_stream_type == 1 or stream.pull_stream_type == 2 or
                            stream.pull_stream_type == 3 or stream.pull_stream_type == 4):
                        if pull_stream_url == "":
                            raise Exception(LANG_VIEWS_T(request, "stream_url_required"))

                        # 检测是否修改摄像头视频地址 start
                        if stream.pull_stream_url != pull_stream_url:
                            # 拉流地址不一致，则首先需要停止拉流代理
                            __del_ret, __del_msg = GlobalUtils.delStreamProxy(stream, lang=lang)
                            stream.forward_state = 0
                            stream.pull_stream_url = pull_stream_url
                        # 检测是否修改摄像头视频地址 end

                    if stream.pull_stream_type != 21:
                        # 非gb28181接入，需要设置分组编号
                        stream.camera_device_id = camera_device_id

                    stream.pull_stream_ip = pull_stream_ip  # v4.405新增
                    stream.pull_stream_port = pull_stream_port  # v4.405新增
                    stream.is_audio = is_audio
                    stream.nickname = nickname
                    stream.remark = remark
                    stream.camera_name = camera_name
                    stream.camera_manufacturer = camera_manufacturer

                    old_record = stream.record_enable
                    stream.record_enable = record_enable

                    stream.last_update_time = datetime.now()
                    stream.save()

                    from app.views.NvrView import _sync_stream_recording
                    _sync_stream_recording(stream)

                    # 添加日志
                    user_id = f_sessionReadUserId(request)
                    lang = f_parseRequestLang(request)
                    LogUtils.add_stream_log(user_id=user_id, stream_code=code,
                                            log_type=LogUtils.LOG_TYPE_EDIT, lang=lang)

                    __msg = LANG_VIEWS_T(request, "msg_edit_success")
                    __ret = True
                else:
                    __msg = LANG_VIEWS_T(request, "stream_not_found")
            except Exception as e:
                __msg = str(e)
        else:
            __msg = __check_msg
    else:
        __msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if __ret else 0,
        "msg": __msg
    }
    g_logger.info("StreamView.openEdit() res:%s" % str(res))
    return f_responseJson(res)
def api_openPlayer(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    info = {}
    if request.method == 'GET':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parseGetParams(request)
            g_logger.info("StreamView.openPlayer() params:%s" % str(params))
            request_ip = f_parseRequestIp(request)  # （v4.725新增） 获取实时请求ip

            app = params.get("app","").strip()
            name = params.get("name","").strip()
            try:
                if app and name:
                    media_info = g_zlm.getMediaInfo(app=app, name=name)
                    media_is_online = 1 if media_info else 0
                    # v4.706确认，返回结构需要保持当前现状
                    stream_info = {
                        "is_online": media_is_online,
                        "app": app,
                        # "produce_speed": produce_speed,
                        # "video": video_str,
                        "aliveSecond": media_info.get("aliveSecond", 0),
                        "totalReaderCount": media_info.get("totalReaderCount", 0),
                        "video_codec_name": media_info.get("video_codec_name", ""),
                        "video_width": media_info.get("video_width", 0),
                        "video_height": media_info.get("video_height", 0),
                        # "audio": audio_str,
                        # "originUrl": d.get("originUrl"),  # 推流地址
                        # "originType": d.get("originType"),  # 推流地址采用的推流协议类型
                        # "originTypeStr": d.get("originTypeStr"),  # 推流地址采用的推流协议类型（字符串）
                        # "clients": d.get("totalReaderCount"),  # 客户端总数量
                        # "schemas_clients": schemas_clients,
                        "videoUrl": g_zlm.get_wsMp4Url(app=app, name=name, request_ip=request_ip),
                        "wsHost": g_zlm.get_wsHost(request_ip=request_ip),
                        "wsMp4Url": g_zlm.get_wsMp4Url(app=app, name=name, request_ip=request_ip),
                        "wsFlvUrl": g_zlm.get_wsFlvUrl(app=app, name=name, request_ip=request_ip),
                        "httpMp4Url": g_zlm.get_httpMp4Url(app=app, name=name, request_ip=request_ip),
                        "httpFlvUrl": g_zlm.get_httpFlvUrl(app=app, name=name, request_ip=request_ip),
                        "rtspUrl": g_zlm.get_rtspUrl(app=app, name=name, request_ip=request_ip),
                        "rtmpUrl": g_zlm.get_rtmpUrl(app=app, name=name, request_ip=request_ip)
                    }
                    stream = StreamModel.objects.filter(app=app, name=name).first()
                    if stream:
                        stream_info["code"] = stream.code
                        stream_info["pull_stream_type"] = stream.pull_stream_type
                    else:
                        stream_info["code"] = name
                        stream_info["pull_stream_type"] = 0
                else:
                    stream_info = {
                        "pull_stream_type": 0,
                        "is_online": 0,
                    }

            except Exception as e:
                g_logger.error("StreamView.openPlayer() e:%s" % str(e))
                stream_info = {
                    "pull_stream_type": 0,
                    "is_online": 0,
                }

            info["stream"] = stream_info
            ret = True
            msg = LANG_VIEWS_T(request, "msg_success")

        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg,
        "info": info
    }
    g_logger.info("StreamView.openPlayer() res:%s" % str(res))
    return f_responseJson(res)

def player(request):
    context = {}
    request_ip = f_parseRequestIp(request)  # （v4.725新增） 获取实时请求ip
    params = f_parseGetParams(request)
    g_logger.info("StreamView.player() request_ip:%s,params:%s" % (request_ip,str(params)))

    app = params.get("app","").strip()
    name = params.get("name","").strip()

    try:
        if app and name:
            media_info = g_zlm.getMediaInfo(app=app, name=name)
            media_is_online = 1 if media_info else 0
            # v4.706确认，返回结构需要保持当前现状
            stream_info = {
                "is_online": media_is_online,
                "app": app,
                "name": name,
                "aliveSecond": media_info.get("aliveSecond", 0),
                "totalReaderCount": media_info.get("totalReaderCount", 0),
                "video_codec_name": media_info.get("video_codec_name", ""),
                "video_width": media_info.get("video_width", 0),
                "video_height": media_info.get("video_height", 0),
                "videoUrl": g_zlm.get_wsMp4Url(app=app, name=name, request_ip=request_ip),
                "wsHost": g_zlm.get_wsHost(request_ip=request_ip),
                "wsMp4Url": g_zlm.get_wsMp4Url(app=app, name=name, request_ip=request_ip),
                "wsFlvUrl": g_zlm.get_wsFlvUrl(app=app, name=name, request_ip=request_ip),
                "httpMp4Url": g_zlm.get_httpMp4Url(app=app, name=name, request_ip=request_ip),
                "httpFlvUrl": g_zlm.get_httpFlvUrl(app=app, name=name, request_ip=request_ip),
                "rtspUrl": g_zlm.get_rtspUrl(app=app, name=name, request_ip=request_ip),
                "rtmpUrl": g_zlm.get_rtmpUrl(app=app, name=name, request_ip=request_ip)
            }

            # 优先：直接按 app+name 查找（live 原始流、rtp 等内部流）
            stream = StreamModel.objects.filter(app=app, name=name).first()
            if stream:
                stream_info["code"] = stream.code
                stream_info["pull_stream_type"] = stream.pull_stream_type
            else:
                stream_info["code"] = name
                stream_info["pull_stream_type"] = 0
        else:
            stream_info = {
                "pull_stream_type": 0,
                "is_online": 0,
            }
    except Exception as e:
        g_logger.error("StreamView.player() e:%s" % str(e))
        stream_info = {
            "pull_stream_type": 0,
            "is_online": 0,
        }

    # 用 json.dumps 序列化为合法 JSON 字符串（True/False -> true/false），
    # 避免 Django 模板 {{ stream|safe }} 把 Python bool 渲染成 JS 标识符导致 "True/False is not defined"
    context["stream_json"] = json.dumps(stream_info, ensure_ascii=False)
    g_logger.info("StreamView.player() stream_info:%s" % str(stream_info))
    return render(request, 'app/stream/player.html', context)

def api_openIndex(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    data = []
    pageData = {}

    if request.method == 'GET':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parseGetParams(request)
            # g_logger.info("StreamView.openIndex() params:%s" % str(params))

            # 同步数据库和在线流状态 start
            medias = g_zlm.getMediaList()
            if len(medias) == 0:
                # 流媒体服务不在线，全部更新下线状态
                g_database.execute("update av_stream set forward_state=0")
            else:
                media_dict = {}
                for m in medias:
                    media_dict[m["an"]] = m

                db_streams = f_dbReadStreamData()
                for d in db_streams:
                    d_id = int(d.get("id", 0))
                    d_forward_state = d.get("forward_state", 0)
                    app_name = "{app}_{name}".format(app=d["app"], name=d["name"])

                    if media_dict.get(app_name):
                        if d_forward_state != 1:
                            g_database.execute("update av_stream set forward_state=1 where id=%d" % d_id)
                    else:
                        if d_forward_state != 0:
                            g_database.execute("update av_stream set forward_state=0 where id=%d" % d_id)
            # 同步数据库和在线流状态 end

            page = params.get('p', 1)
            page_size = params.get('ps', 10)
            search_text = str(params.get('search_text', "")).strip()  # v4.638新增
            search_status = int(params.get('search_status', -1))   # v4.716新增 转发状态筛选：-1全部 0未转发 1转发中

            try:
                page = int(page)
            except:
                page = 1

            try:
                page_size = int(page_size)
                if page_size < 1:
                    page_size = 1
            except:
                page_size = 10

            # 构建查询条件
            conditions = []
            if search_text:
                conditions.append("nickname like '%{search_text}%' or code = '{search_text}' ".format(search_text=search_text))
            
            if search_status != -1:
                conditions.append("forward_state = {status}".format(status=search_status))
            
            # 拼接SQL
            if len(conditions) > 0:
                where_condition = " where " + " and ".join(conditions)
                sql = "select * from av_stream {where_condition} order by id desc".format(where_condition=where_condition)
            else:
                sql = "select * from av_stream order by id desc"
            
            skip = (page - 1) * page_size

            data = g_database.select(sql)
            rec_ids = set()
            try:
                from app.recording.manager import get_recording_manager
                rec_ids = set(get_recording_manager().list_recording())
            except Exception:
                pass
            for d in data:
                camera_device_id = d.get("camera_device_id")
                if camera_device_id is None or camera_device_id == "":
                    camera_device_id = d["code"]
                d["camera_device_id"] = camera_device_id
                sid = int(d.get("id", 0) or 0)
                d["record_enable"] = int(d.get("record_enable", 0) or 0)
                d["is_recording"] = 1 if sid in rec_ids else 0

            grouped_data = group_by_field(data, field="camera_device_id")
            count = len(grouped_data)
            if count > 0:
                grouped_data = grouped_data[skip:skip + page_size]
                for i, group in enumerate(grouped_data):
                    # print(f"组 {i + 1}: {group}")

                    for d in group:
                        camera_name = d.get("camera_name","unknown name")
                        d["camera_name"] = camera_name
                        camera_manufacturer = d.get("camera_manufacturer","unknown manufacturer")
                        d["camera_manufacturer"] = camera_manufacturer
                        d["last_update_time"] = d.get("last_update_time").strftime("%Y/%m/%d %H:%M")

                data = grouped_data
                ret = True
                msg = LANG_VIEWS_T(request, "msg_success")

            else:
                data = []
                msg = LANG_VIEWS_T(request, "msg_empty")

            page_num = int(count / page_size)  # 总页数
            if count % page_size > 0:
                page_num += 1

            pageData = {
                "page": page,
                "page_size": page_size,
                "page_num": page_num,
                "count": count,
                "pageLabels": buildPageLabels(page=page, page_num=page_num, lang=f_parseRequestLang(request))
            }
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg,
        "data": data,
        "pageData": pageData,
        "extra": {
            "audioTypes": get_audio_types(f_parseRequestLang(request)),
            "pullStreamTypes": g_pull_stream_types
        }
    }
    return f_responseJson(res)

def api_openExportFile(request):
    """导出所有摄像头数据为Excel文件（与导入模板格式一致）"""
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")

    if request.method == 'GET':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            g_logger.info("StreamView.openExportFile()")

            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            # lang = f_parseRequestLang(request)

            try:
                # 查询所有摄像头数据
                streams = StreamModel.objects.all().order_by("id")

                # 创建工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = "摄像头数据"

                # 定义表头（与导入模板完全一致）
                headers = [
                    "摄像头编号", "摄像头昵称", "拉流地址", "IP地址", "端口",
                    "用户名", "密码", "备注", "是否音频", "分组编号"
                ]

                # 表头样式
                header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                # 写入表头
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border

                # 数据样式
                data_font = Font(name='微软雅黑', size=10)
                data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

                # 写入数据行
                for row_idx, stream in enumerate(streams, 2):
                    # 设备ID处理：如果为空则使用code
                    camera_device_id = stream.camera_device_id if stream.camera_device_id else stream.code

                    row_data = [
                        stream.code or "",
                        stream.nickname or "",
                        stream.pull_stream_url or "",
                        stream.pull_stream_ip or "",
                        stream.pull_stream_port or 554,
                        stream.pull_stream_username or "",
                        stream.pull_stream_password or "",
                        stream.remark or "",
                        stream.is_audio or 0,
                        camera_device_id
                    ]

                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = data_font
                        cell.alignment = data_align
                        cell.border = thin_border

                # 设置列宽
                col_widths = [20, 20, 40, 18, 10, 18, 18, 30, 10, 20]
                for col_idx, width in enumerate(col_widths, 1):
                    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

                # 设置行高（表头）
                ws.row_dimensions[1].height = 25

                # 生成响应
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                response = HttpResponse(buffer.getvalue(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = "cameras_export_%s.xlsx" % datetime.now().strftime("%Y%m%d_%H%M%S")
                response['Content-Disposition'] = 'attachment; filename="%s"' % filename

                return response

            except Exception as e:
                msg = str(e)
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg
    }
    g_logger.info("StreamView.openExportFile() res:%s" % str(res))

    return f_responseJson(res)

def api_openImportFile(request):
    __ret = False
    __msg = LANG_VIEWS_T(request, "msg_unknown_error")
    if request.method == 'POST':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parsePostParams(request)
            lang = f_parseRequestLang(request)
            g_logger.info("StreamView.openImportFile() params:%s" % str(params))
            try:
                file = request.FILES.get("file")
                importRemark = str(params.get("importRemark", "")).strip()

                upload_utils = UploadUtils()
                __upload_ret, __upload_msg, __upload_data = upload_utils.upload_camera_xlsx(file=file,
                                                                                            upload_dir=g_config.storageTempDir,
                                                                                            lang=lang)

                if __upload_ret:
                    user_id = f_sessionReadUserId(request)
                    success_count = 0
                    error_count = 0
                    for d in __upload_data:
                        try:
                            stream_code = d["code"]
                            if not stream_code:  # 未填写编号，则自动生成编号
                                stream_code = GB28181CodeUtils().generate_by_time()

                            nickname = d["nickname"]
                            pull_stream_url = d["pull_stream_url"]
                            pull_stream_ip = d["pull_stream_ip"]
                            pull_stream_port = d["pull_stream_port"]
                            username = d["username"]
                            password = d["password"]
                            camera_device_id = str(d.get("camera_device_id", "")).strip()
                            is_audio = int(d.get("is_audio", 0))
                            remark = str(d.get("remark", "")).strip()

                            remark = "remark=%s,importRemark=%s" % (remark, importRemark)

                            if camera_device_id is None or camera_device_id == "":
                                camera_device_id = stream_code

                            if pull_stream_url.startswith("rtsp://") or pull_stream_url.startswith("rtsps://"):
                                pull_stream_type = 1
                            elif pull_stream_url.startswith("rtmp://") or pull_stream_url.startswith("rtmps://"):
                                pull_stream_type = 2
                            elif pull_stream_url.startswith("http://") or pull_stream_url.startswith("https://"):
                                pull_stream_type = 3
                            else:
                                raise Exception(LANG_VIEWS_T(request, "stream_url_format_error"))

                            stream = StreamModel.objects.filter(code=stream_code).first()
                            if stream:
                                raise Exception(LANG_VIEWS_T(request, "stream_code_exists"))
                            else:
                                stream = StreamModel()
                                stream.user_id = user_id
                                stream.sort = 0
                                stream.code = stream_code
                                stream.app = g_zlm.default_stream_app
                                stream.name = stream_code
                                stream.pull_stream_url = pull_stream_url
                                stream.pull_stream_type = pull_stream_type  # v4.405新增
                                stream.pull_stream_transfer_mode = 0  # v4.405新增
                                stream.pull_stream_ip = pull_stream_ip  # v4.405新增
                                stream.pull_stream_port = pull_stream_port  # v4.405新增
                                stream.pull_stream_username = username
                                stream.pull_stream_password = password
                                stream.nickname = nickname
                                stream.remark = remark
                                stream.forward_state = 0  # 默认未开启转发
                                stream.camera_device_id = camera_device_id # v4.638新增

                                stream.create_time = datetime.now()
                                stream.last_update_time = datetime.now()
                                stream.add_type = 1
                                stream.state = 0
                                stream.is_audio = is_audio  # v4.405新增

                                stream.save()

                                success_count += 1
                        except Exception as e:
                            g_logger.error("StreamView.openImportFile() error: %s, d=%s" % (str(e),str(d)))
                            error_count += 1

                    __ret = True
                    __msg = LANG_VIEWS_T(request, "msg_batch_result") % (success_count, error_count)

                else:
                    __msg = __upload_msg
            except Exception as e:
                __msg = str(e)
        else:
            __msg = __check_msg
    else:
        __msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if __ret else 0,
        "msg": __msg
    }
    g_logger.info("StreamView.openImportFile() res:%s" % str(res))

    return f_responseJson(res)

def api_openDel(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    lang = f_parseRequestLang(request)
    if request.method == 'POST':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parsePostParams(request)
            g_logger.info("StreamView.openDel() params:%s" % str(params))

            handle = params.get("handle") # one：删除一个视频流，all：删除全部视频流
            stream_code = params.get("code")

            if handle == "one":
                stream = StreamModel.objects.filter(code=stream_code).first()
                if stream:
                    __ret, __msg = GlobalUtils.delStreamProxy(stream, lang=lang)
                    if stream.delete():
                        ret = True
                        msg = LANG_VIEWS_T(request, "msg_success")
                    else:
                        msg = LANG_VIEWS_T(request, "stream_delete_failed")
                else:
                    msg = LANG_VIEWS_T(request, "stream_not_found")
            elif handle == "all":
                success_count = 0
                error_count = 0
                streams = StreamModel.objects.all()
                for stream in streams:
                    __ret, __msg = GlobalUtils.delStreamProxy(stream, lang=lang)
                    if stream.delete():
                        g_gb28181SipServer.remove_channel(stream.code)
                        success_count += 1
                    else:
                        error_count += 1
                ret = True
                msg = LANG_VIEWS_T(request, "msg_batch_result") % (success_count, error_count)

            else:
                msg = LANG_VIEWS_T(request, "msg_invalid_parameter")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg
    }
    g_logger.info("StreamView.openDel() res:%s" % str(res))
    return f_responseJson(res)

def api_openHandleAllStreamProxy(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    lang = f_parseRequestLang(request)

    if request.method == 'POST':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parsePostParams(request)
            g_logger.info("StreamView.openHandleAllStreamProxy() params:%s" % str(params))

            handle = params.get("handle")  # add,del (add 开启流，del 关闭流)

            if handle == "add":  # 全部开启转发
                ret,msg = GlobalUtils.addAllStreamProxy(lang=lang)
            elif handle == "del":  # 全部关闭转发
                GlobalUtils.delAllStreamProxy(lang=lang)
                ret = True
                msg = LANG_VIEWS_T(request, "msg_success")
            else:
                msg = LANG_VIEWS_T(request, "msg_invalid_parameter")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg
    }
    g_logger.info("StreamView.openHandleAllStreamProxy() res:%s" % str(res))
    return f_responseJson(res)


def api_openAddStreamProxy(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    lang = f_parseRequestLang(request)
    if request.method == 'POST':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parsePostParams(request)
            g_logger.info("StreamView.openAddStreamProxy() params:%s" % str(params))
            
            stream_code = params.get("code", "").strip()
            stream = StreamModel.objects.filter(code=stream_code).first()
            if stream:
                if stream.forward_state == 1:
                    ret = True
                    msg = LANG_VIEWS_T(request, "stream_forward_already_enabled")
                else:
                    ret, msg = GlobalUtils.addStreamProxy(stream, lang=lang)
                    if ret:
                        stream.forward_state = 1
                        stream.save()
                        from app.views.NvrView import _sync_stream_recording
                        _sync_stream_recording(stream)
                        ret = True
                        msg = LANG_VIEWS_T(request, "stream_forward_enabled_success")
            else:
                msg = LANG_VIEWS_T(request, "stream_not_found")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg
    }
    g_logger.info("StreamView.openAddStreamProxy() res:%s" % str(res))
    return f_responseJson(res)
def api_openDelStreamProxy(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    lang = f_parseRequestLang(request)
    if request.method == 'POST':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parsePostParams(request)
            g_logger.info("StreamView.openDelStreamProxy() params:%s" % str(params))
            
            stream_code = params.get("code", "").strip()
            stream = StreamModel.objects.filter(code=stream_code).first()
            if stream:
                __ret, __msg = GlobalUtils.delStreamProxy(stream, lang=lang)
                stream.forward_state = 0
                stream.save()
                from app.views.NvrView import _sync_stream_recording
                _sync_stream_recording(stream)

                ret = __ret
                msg = __msg
            else:
                msg = LANG_VIEWS_T(request, "stream_not_found")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg
    }
    g_logger.info("StreamView.openDelStreamProxy() res:%s" % str(res))
    return f_responseJson(res)
def api_openPtz(request):
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    lang = f_parseRequestLang(request)
    if request.method == 'POST':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parsePostParams(request)
            g_logger.info("StreamView.openPtz() params:%s" % str(params))

            code = params.get("code", "").strip()
            ptzType = int(params.get("ptzType", 0))
            val = int(params.get("val", 0))

            if ptzType < 0 or ptzType > 11:
                raise Exception(LANG_VIEWS_T(request, "ptz_type_range_error"))

            stream = StreamModel.objects.filter(code=code).first()
            if stream:
                if stream.pull_stream_type == 21:

                    ret, msg = g_gb28181SipServer.request_ptz(client_id=stream.camera_device_id,
                                                   channel_id=stream.code,
                                                   ptz_type=ptzType,
                                                   val=val)
                else:
                    msg = LANG_VIEWS_T(request, "stream_pull_type_not_supported")
            else:
                msg = LANG_VIEWS_T(request, "stream_not_found")
        else:
            msg  = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg
    }
    g_logger.info("StreamView.openPtz() res:%s" % str(res))
    return f_responseJson(res)


def api_openGetAllStreamData(request):
    """在线预览页面获取所有在线流数据（原 OpenView 中的同名接口迁移至此）"""
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    data = []
    request_ip = f_parseRequestIp(request)

    if request.method == 'GET':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            params = f_parseGetParams(request)
            handle = params.get("handle", "").strip()
            g_logger.info("StreamView.openGetAllStreamData() request_ip:%s,params:%s" % (request_ip, str(params)))

            # 获取所有 ZLM 在线流
            medias = g_zlm.getMediaList(request_ip=request_ip)

            # 构建 app_name -> 摄像头信息 映射（从数据库）
            db_streams = g_database.select("select id,app,name,code,nickname from av_stream")
            db_dict = {}
            for s in db_streams:
                key = "{app}_{name}".format(app=s["app"], name=s["name"])
                db_dict[key] = s

            for m in medias:
                app = m.get("app", "")
                name = m.get("name", "")
                app_name = m.get("app_name", "{0}_{1}".format(app, name))

                nickname = name
                code = app_name

                db_info = db_dict.get(app_name)
                if db_info:
                    code = db_info.get("code") or app_name
                    nickname = db_info.get("nickname") or name
                    m["id"] = db_info.get("id")

                m["code"] = code
                m["nickname"] = nickname
                data.append(m)

            ret = True
            msg = LANG_VIEWS_T(request, "msg_success")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg,
        "data": data
    }
    g_logger.info("StreamView.openGetAllStreamData() res.data length:%d" % len(data))
    return f_responseJson(res)


def api_openGetStatisticsStream(request):
    """首页流媒体统计（原 OpenView 接口迁移至此）"""
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    info = {"groupCount": 0, "count": 0, "forwardCount": 0}
    request_ip = f_parseRequestIp(request)

    if request.method == 'GET':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            g_logger.info("StreamView.openGetStatisticsStream() request_ip:%s" % request_ip)
            try:
                # 接入数量 = 数据库中摄像头总数
                db_rows = g_database.select("select app,name from av_stream")
                info["count"] = len(db_rows)

                # 在线数量 = ZLM 中实际存在的流数量
                medias = g_zlm.getMediaList(request_ip=request_ip)
                info["forwardCount"] = len(medias)

                # 分组数量 = 不同 app 的数量
                apps = set()
                for r in db_rows:
                    if r.get("app"):
                        apps.add(r["app"])
                info["groupCount"] = len(apps)
            except Exception as e:
                g_logger.error("StreamView.openGetStatisticsStream() e:%s" % str(e))

            ret = True
            msg = LANG_VIEWS_T(request, "msg_success")
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")

    res = {
        "code": 1000 if ret else 0,
        "msg": msg,
        "info": info
    }
    g_logger.info("StreamView.openGetStatisticsStream() res:%s" % str(res))
    return f_responseJson(res)


def api_openOnvifDiscover(request):
    """WS-Discovery 发现局域网 ONVIF 设备"""
    ret = False
    msg = LANG_VIEWS_T(request, "msg_unknown_error")
    data = []
    if request.method == 'GET':
        __check_ret, __check_msg = f_checkRequestSafe(request)
        if __check_ret:
            try:
                from app.services.onvif_discovery import discover_onvif, get_rtsp_url_from_onvif
                params = f_parseGetParams(request)
                timeout = float(params.get("timeout", 3))
                username = (params.get("username") or "").strip()
                password = (params.get("password") or "").strip()
                resolve = str(params.get("resolve_rtsp", "0")).strip() in ("1", "true", "yes")
                devices = discover_onvif(timeout=timeout)
                for d in devices:
                    item = dict(d)
                    if resolve and d.get("xaddr"):
                        rtsp = get_rtsp_url_from_onvif(d["xaddr"], username, password)
                        if rtsp:
                            item["rtsp_url"] = rtsp
                    data.append(item)
                ret = True
                msg = LANG_VIEWS_T(request, "msg_success")
            except Exception as e:
                msg = str(e)
                g_logger.exception("openOnvifDiscover: %s", e)
        else:
            msg = __check_msg
    else:
        msg = LANG_VIEWS_T(request, "msg_method_not_supported")
    return f_responseJson({"code": 1000 if ret else 0, "msg": msg, "data": data})